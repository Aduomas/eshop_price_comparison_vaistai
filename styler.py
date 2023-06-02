import pandas as pd
from openpyxl import load_workbook
import os
from sku_obtainer import VaistaiCrawler
import asyncio
from olap_db import fetch_dataframe
from utils import get_herba_price


def calculate_change(row):
    min_index = row.idxmin()
    if min_index == "herba":
        second_min = row[row != row[min_index]].min()
        return ((row["herba"] - second_min) / row["herba"] * 100).round(2) * -1
    else:
        return ((row["herba"] - row[min_index]) / row["herba"] * 100).round(2) * -1


def color_cells(row):
    green = "background-color: green"
    white = ""  # default color
    minimum_value = row.drop(["SKU", "Revenue", "Change (%)"]).min()
    colors = [green if cell == minimum_value else white for cell in row]
    return colors


def adjust_column_width(workbook_path):
    wb = load_workbook(workbook_path)
    for sheet in wb.sheetnames:  # go through all sheets
        worksheet = wb[sheet]
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            if column[0].column_letter == "A":
                adjusted_width -= 20
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    wb.save(workbook_path)


async def get_sku_list(title_list):
    crawler = VaistaiCrawler("sksherba", "sksherba", async_req=True)
    await crawler.authenticate()
    results = await crawler.search(title_list)
    results = [r[0] if r else "Not found" for r in results]
    return results


def process_data(df, file_name):

    reshaped_df = df.pivot(index="title", columns="eshop", values="price")

    reshaped_df = reshaped_df.rename(columns={"tamro": "benu", "limedika": "gintarine"})
    reshaped_df["Change (%)"] = reshaped_df.apply(calculate_change, axis=1)
    reshaped_df["Change (%)"] = reshaped_df["Change (%)"].apply(
        lambda x: f"{x}%" if pd.notna(x) else ""
    )

    reshaped_df["SKU"] = asyncio.run(get_sku_list(reshaped_df.index.tolist()))

    revenue_df = fetch_dataframe()
    mapping = revenue_df.set_index("Kodas")["BePVM"].astype(int)

    reshaped_df["Revenue"] = reshaped_df["SKU"].map(mapping)
    reshaped_df.sort_values(by="Revenue", ascending=False, inplace=True)
    reshaped_df["Revenue"] = reshaped_df["Revenue"].astype(str) + " €"

    reshaped_df.loc[:, "herba"] = reshaped_df["SKU"].apply(
        lambda x: get_herba_price(f"https://www.herba.lt/{x}")
        if x != "Not found"
        else None
    )

    cols = reshaped_df.columns.tolist()
    desired_order = [
        "SKU",
        "Revenue",
        "Change (%)",
        "herba",
        "eurovaistine",
        "benu",
        "gintarine",
        "camelia",
    ]

    for eshop in reversed(
        desired_order
    ):  # reverse order to preserve the order in the desired_order list
        try:
            cols.insert(0, cols.pop(cols.index(eshop)))
        except ValueError:
            pass

    reshaped_df = reshaped_df.reindex(columns=cols)
    styled_df = reshaped_df.style.apply(color_cells, axis=1)

    # Write the DataFrame to Excel
    file_path = f"{file_name}.xlsx"
    styled_df.to_excel(file_path, engine="openpyxl")

    adjust_column_width(file_path)
